"""
Flora Gu
2023-06-05
"""

#pip install PyPDF2
#pip install openpyxl
from openpyxl import Workbook
import os
from PyPDF2 import PdfReader

def keep(item, text):
    """sig: str, str-> list
    change the string text into a list which separates according to \n
    and keep any thing in the list after the string item"""
    text=text.split('\n')
    k=''
    new=[]
    for i in range(0, len(text),1):
        print (text[i])
        if text[i]==item:
            k=i
    for i in range(k+1, len(text)):
        new.append(text[i])
    return new
           

def keep2(text1):
    """sig: str->list
    remove all numbers and symbol - from the string
    and then change string into lists with different organizations"""
    for letter in text1:
        if letter.isdigit():
            text1=text1.replace(letter," ")
        elif letter=='-':
            text1=text1.replace(letter," ") #remove all numbers and symbol -

    text1=text1.split(" ")
    for word in text1:
        for i in range(0, len(word)):
            if word[i]==" ":
                word=word.replace(word[i],"") # change str into list and removes all empty spaces in each entry of the list

    while '' in text1: #removes all empty entries in the list
        text1.remove('') 
            
    return text1


def create_dict(purchaser, organization):
    """sig:str, list->dict
    create a dictionary that matches the organization name with purchaser
    the key is the name of organization and value is the percentage of match"""
    dict1={}
    for each in organization:
        count=0
        for i in range(0,len(each)):
            if each[i] in purchaser:
                count += 1
        percentage=count/len(each)
        dict1[each]=percentage
    return dict1


def rev_look_biggest(mydict):
    """sig: dict->list
    find the biggest percentage in the values of the dictionary
    and find the keys to the values.
    Return a dictionary with key of the percentage and value of the matched list"""

    values=mydict.values()
    big=0
    for value in values:
        if value>big:
            big=value    
    
    match_key=[]
    for key in mydict:
        if mydict[key]==big:
            match_key.append(key)
    newdict={}
    newdict[big]=match_key
    return newdict


#change to your own directory
os.chdir("C:/Users/flora/Dropbox/RA- Flora Gu/Data")

file=open("中央组织_手动整理.txt", "r", encoding="utf-8")
organization=[]
for line in file:
    line=line[0:len(line)-1] #the last letter is \n.
    organization.append(line)
file.close()

print('predone')

from openpyxl import load_workbook
workbook = load_workbook(filename="政府采购合同公告数据_raw.xlsx")
sheet = workbook.active

print('First step done')

for i in range(2,sheet.max_row+1): 
    ce='I'+str(i)
    cellvalue=str(sheet[ce].value)
    newcell='U'+str(i)
    newcell2='V'+str(i)
    org=[]
    for each in organization:
        if each in cellvalue:
            sheet[newcell].value=1
            org.append(each)
    if sheet[newcell].value!=1:
        sheet[newcell].value=0
    sheet[newcell2].value=str(org)

print('Second step done')

    
workbook.save(filename="政府采购合同公告数据_python_edited.xlsx")
print('Final step done')


